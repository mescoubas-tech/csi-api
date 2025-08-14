# app/services/schedule_checker.py
from __future__ import annotations
from pathlib import Path
from typing import Iterable, List, Dict, Tuple
from datetime import datetime, timedelta, date
import csv

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None

from ..core.config import get_settings
from ..models.schemas import SchedulesCheckResult, ScheduleViolation, ScheduleStat
from .pdf_schedule_parser import parse_pdf_schedules

TIME_FMT = "%H:%M"

ALIASES = {
    "agent_id": {"agent_id", "agent", "matricule", "id"},
    "date": {"date", "jour"},
    "start_time": {"start_time", "start", "debut", "début", "heure_debut", "debut_poste"},
    "end_time": {"end_time", "end", "fin", "heure_fin", "fin_poste"},
    "break_minutes": {"break_minutes", "break", "pause", "pause_min", "pause (min)"},
}

def _norm_header(name: str) -> str:
    n = name.strip().lower().replace("é","e").replace("è","e").replace("ê","e").replace("’","'")
    for canon, al in ALIASES.items():
        if n in al:
            return canon
    return name

def _parse_time(s: str) -> datetime:
    s = s.strip().lower().replace(" ", "").replace("h", ":").replace(".", ":")
    hh, mm = s.split(":")
    return datetime.strptime(f"{int(hh):02d}:{int(mm):02d}", TIME_FMT)

def _parse_date(s: str) -> date:
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except Exception:
            pass
    s2 = s.replace(".", "/").replace("-", "/")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s2.strip(), fmt).date()
        except Exception:
            pass
    raise ValueError(f"Date invalide: {s}")

def _dur_minutes(start: datetime, end: datetime) -> int:
    # gestion nuit: fin < début => +24h
    if end <= start:
        end = end + timedelta(days=1)
    return int((end - start).total_seconds() // 60)

def _read_csv(p: Path) -> List[Dict]:
    rows: List[Dict] = []
    with open(p, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        header = {k: _norm_header(k) for k in reader.fieldnames or []}
        for r in reader:
            rec = {header.get(k, k): (r[k] if r.get(k) is not None else "") for k in r}
            rows.append(rec)
    return rows

def _read_xlsx(p: Path) -> List[Dict]:
    if openpyxl is None:
        return []
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    header_map: Dict[int, str] = {}
    rows: List[Dict] = []
    for j, cell in enumerate(ws[1], 1):
        if cell.value:
            header_map[j] = _norm_header(str(cell.value))
    for i in range(2, ws.max_row + 1):
        rec: Dict[str, str] = {}
        for j in range(1, ws.max_column + 1):
            key = header_map.get(j)
            if not key:
                continue
            val = ws.cell(row=i, column=j).value
            rec[key] = str(val) if val is not None else ""
        rows.append(rec)
    return rows

def _normalize_rows(raw: List[Dict]) -> List[Dict]:
    out: List[Dict] = []
    for r in raw:
        try:
            agent = (r.get("agent_id") or "").strip()
            if not agent:
                continue
            d = _parse_date(r.get("date",""))
            st = _parse_time(str(r.get("start_time","")))
            en = _parse_time(str(r.get("end_time","")))
            br = int(r.get("break_minutes") or 0)
            out.append({
                "agent_id": agent,
                "date": d,
                "start": st,
                "end": en,
                "break_min": br
            })
        except Exception:
            continue
    return out

def _group_by_agent(rows: List[Dict]) -> Dict[str, List[Dict]]:
    g: Dict[str, List[Dict]] = {}
    for r in rows:
        g.setdefault(r["agent_id"], []).append(r)
    for a in g:
        g[a].sort(key=lambda x: (x["date"], x["start"].time()))
    return g

def check_schedules(paths: Iterable[Path]) -> SchedulesCheckResult:
    paths = [Path(p) for p in paths]
    raw: List[Dict] = []

    # PDF
    pdf_rows = parse_pdf_schedules([p for p in paths if p.suffix.lower() == ".pdf"])
    raw.extend(pdf_rows)

    # CSV / XLSX
    for p in paths:
        suf = p.suffix.lower()
        if suf == ".csv":
            raw.extend(_read_csv(p))
        elif suf in (".xlsx", ".xlsm"):
            raw.extend(_read_xlsx(p))

    rows = _normalize_rows(raw)
    groups = _group_by_agent(rows)

    S = get_settings()
    violations: List[ScheduleViolation] = []
    stats: List[ScheduleStat] = []

    for agent, shifts in groups.items():
        # calculs par jour & semaine
        daily_minutes: Dict[date, int] = {}
        weeks: Dict[Tuple[int,int], int] = {}  # (year, iso_week) -> minutes
        last_shift_end_by_day: Dict[date, datetime] = {}
        # pour repos quotidien/harmonisation, on garde une timeline
        previous_end: datetime | None = None
        consec_days = 0
        last_day: date | None = None

        for sh in shifts:
            minutes = _dur_minutes(sh["start"], sh["end"]) - int(sh["break_min"] or 0)
            minutes = max(0, minutes)
            d = sh["date"]
            daily_minutes[d] = daily_minutes.get(d, 0) + minutes

            iso = d.isocalendar()  # year, week, weekday
            key = (iso[0], iso[1])
            weeks[key] = weeks.get(key, 0) + minutes

            # Repos quotidien (11h)
            if previous_end is not None:
                rest = (sh["start"] - previous_end).total_seconds() / 3600.0
                if rest < S.MIN_DAILY_REST_HOURS:
                    violations.append(ScheduleViolation(
                        agent_id=agent,
                        type="DAILY_REST",
                        date=d.isoformat(),
                        details=f"Repos quotidien {rest:.1f}h < {S.MIN_DAILY_REST_HOURS}h"
                    ))
            previous_end = sh["end"]
            if previous_end <= sh["start"]:
                previous_end = previous_end + timedelta(days=1)

            # Jours consécutifs
            if last_day is None or (d - last_day).days == 1:
                consec_days += 1
            elif d == last_day:
                pass
            else:
                consec_days = 1
            last_day = d
            if consec_days > S.MAX_CONSECUTIVE_DAYS:
                violations.append(ScheduleViolation(
                    agent_id=agent,
                    type="CONSEC_DAYS",
                    date=d.isoformat(),
                    details=f"{consec_days} jours consécutifs > {S.MAX_CONSECUTIVE_DAYS}"
                ))

        # seuils journaliers
        for d, mins in daily_minutes.items():
            if mins > int(S.MAX_HOURS_PER_DAY * 60):
                violations.append(ScheduleViolation(
                    agent_id=agent,
                    type="DAILY_MAX",
                    date=d.isoformat(),
                    details=f"{mins/60:.2f} h > {S.MAX_HOURS_PER_DAY} h / jour"
                ))

        # seuils hebdomadaires & moyenne 12 semaines
        weeks_sorted = sorted(weeks.items())
        for (y, w), mins in weeks_sorted:
            if mins > int(S.MAX_HOURS_PER_WEEK * 60):
                violations.append(ScheduleViolation(
                    agent_id=agent,
                    type="WEEKLY_MAX",
                    week=f"{y}-W{w:02d}",
                    details=f"{mins/60:.2f} h > {S.MAX_HOURS_PER_WEEK} h / semaine"
                ))

        # moyenne glissante sur 12 semaines
        if len(weeks_sorted) >= 12:
            for i in range(0, len(weeks_sorted) - 11):
                window = weeks_sorted[i:i+12]
                total = sum(m for _, m in window)
                avg_h = (total/12) / 60.0
                if avg_h > S.AVG_HOURS_PER_12W:
                    startw = f"{window[0][0][0]}-W{window[0][0][1]:02d}"
                    endw   = f"{window[-1][0][0]}-W{window[-1][0][1]:02d}"
                    violations.append(ScheduleViolation(
                        agent_id=agent,
                        type="AVG_12W",
                        week=f"{startw}→{endw}",
                        details=f"moyenne {avg_h:.2f} h > {S.AVG_HOURS_PER_12W} h / 12 sem."
                    ))

        # stats
        total_min = sum(daily_minutes.values())
        stats.append(ScheduleStat(
            agent_id=agent,
            total_hours=round(total_min/60.0, 2),
            days_worked=len(daily_minutes),
            weeks_count=len(weeks)
        ))

    return SchedulesCheckResult(
        agents=sorted(set(groups.keys())),
        stats=stats,
        violations=violations,
        extras={}
    )
